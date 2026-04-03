import json
import os
import re
import threading
import concurrent.futures
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from docx import Document
from anthropic import Anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import lxml.etree as etree
import yaml

APP_TITLE               = "Indeed Proto QA Reviewer"
WORD_COST_USD           = 0.005
DEFAULT_CLAUDE_MODEL_AGENT1 = "claude-opus-4-5"
DEFAULT_CLAUDE_MODEL_AGENT2 = "claude-haiku-4-5"
DEFAULT_MOSAIQ_CONFIG   = Path(
    r"C:\Internal Apps\MosAIQ LQA\MosAIQ-LQA\config.yaml"
)
MAX_WORKERS             = 6
FP_CONFIDENCE_THRESHOLD = 0.65

# ── Colour palette ────────────────────────────────────────────────────────────
BG_DARK        = "#1e1f2e"
BG_PANEL       = "#252637"
BG_CARD        = "#2d2f45"
BG_INPUT       = "#1a1b2a"
ACCENT         = "#7c6af7"
ACCENT_HOVER   = "#9b8df9"
ACCENT_DIM     = "#4a4580"
SUCCESS        = "#4caf82"
WARNING        = "#f0a843"
DANGER         = "#e05c6e"
TEXT_PRIMARY   = "#e8e9f3"
TEXT_SECONDARY = "#9091a8"
TEXT_MUTED     = "#5c5d7a"
BORDER         = "#3a3b52"


# ═══════════════════════════════════════════════════════════════════════════════
#  Data classes
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class XliffSegment:
    seg_id:    str
    source:    str
    target:    str
    state:     str
    file_name: str
    doc_id:    str = ""
    seg_seq:   str = ""
    extra:     Dict[str, Any] = field(default_factory=dict)


@dataclass
class MosaiqIssue:
    row_index:   int
    seg_id:      str
    source_text: str
    target_text: str
    error_type:  str
    severity:    str
    comment:     str
    doc_id:      str = ""
    seg_seq:     str = ""
    seg_info:    str = ""
    fp:          str = ""
    issue_sign:  str = ""
    issue_group: str = ""
    extra:       Dict[str, Any] = field(default_factory=dict)


@dataclass
class EvaluatedSegment:
    segment:           XliffSegment
    has_issue:         bool
    is_false_positive: bool
    ai_confidence:     float
    ai_issue:          str
    ai_comment:        str
    ai_rationale:      str
    matched_mosaiq:    Optional[MosaiqIssue] = None


# ═══════════════════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def safe_json_loads(text: str) -> Optional[dict]:
    try:
        return json.loads(text)
    except Exception:
        pass
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            return None
    return None


def load_env_file(path: Path) -> Dict[str, str]:
    env: Dict[str, str] = {}
    if not path.exists():
        return env
    try:
        for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            env[key.strip()] = value.strip().strip('"').strip("'")
    except Exception:
        pass
    return env


def word_count(text: str) -> int:
    return len(re.findall(r"\b\w+\b", text))


def save_json(path: Path, data: dict) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps(data, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception:
        pass


def load_agent_settings(config_path: Path) -> dict:
    settings = {
        "step1_model": DEFAULT_CLAUDE_MODEL_AGENT1,
        "step1_temp":  0.2,
        "step2_model": DEFAULT_CLAUDE_MODEL_AGENT2,
        "step2_temp":  0.2,
    }
    if not config_path.exists():
        return settings
    try:
        data  = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
        lqa   = data.get("lqa", {}) or {}
        step1 = lqa.get("step1", {}) or {}
        step2 = lqa.get("step2", {}) or {}
        settings["step1_model"] = str(step1.get("model") or settings["step1_model"])
        settings["step2_model"] = str(step2.get("model") or settings["step2_model"])
        settings["step1_temp"]  = float(step1.get("temperature", settings["step1_temp"]))
        settings["step2_temp"]  = float(step2.get("temperature", settings["step2_temp"]))
    except Exception:
        pass
    return settings


def instructions_to_json(text: str, source_file: str) -> dict:
    return {"source_file": source_file, "instructions": text}


def segments_to_json(segments: List[XliffSegment]) -> dict:
    return {
        "segments": [
            {
                "seg_id":    s.seg_id,
                "source":    s.source,
                "target":    s.target,
                "state":     s.state,
                "file_name": s.file_name,
                "doc_id":    s.doc_id,
                "seg_seq":   s.seg_seq,
                **s.extra,
            }
            for s in segments
        ]
    }


def _results_to_json_list(results: List[EvaluatedSegment]) -> list:
    """
    Convert evaluated segments to a serialisable list.
    Kept as a standalone function so it is never nested inside a try block.
    """
    rows = []
    for r in results:
        if r.has_issue:
            disposition = "real_error"
        elif r.is_false_positive and r.ai_confidence >= FP_CONFIDENCE_THRESHOLD:
            disposition = "fp_excluded"
        elif r.is_false_positive:
            disposition = "fp_low_conf"
        else:
            disposition = "clean"

        rows.append({
            "seg_id":            r.segment.seg_id,
            "doc_id":            r.segment.doc_id,
            "seg_seq":           r.segment.seg_seq,
            "source":            r.segment.source,
            "target":            r.segment.target,
            "state":             r.segment.state,
            "has_issue":         r.has_issue,
            "is_false_positive": r.is_false_positive,
            "ai_confidence":     r.ai_confidence,
            "ai_issue":          r.ai_issue,
            "ai_comment":        r.ai_comment,
            "ai_rationale":      r.ai_rationale,
            "disposition":       disposition,
        })
    return rows


def _build_file_json(
    ts:           str,
    xp:           Path,
    inst_path:    Path,
    report_path:  Path,
    model:        str,
    results:      List[EvaluatedSegment],
    error_cnt:    int,
    excl_cnt:     int,
    low_conf_cnt: int,
    clean_cnt:    int,
    total_words:  int,
) -> dict:
    return {
        "run_timestamp":      ts,
        "xliff_file":         str(xp),
        "instructions":       str(inst_path),
        "mosaiq_report":      str(report_path),
        "model":              model,
        "workers":            MAX_WORKERS,
        "fp_threshold":       FP_CONFIDENCE_THRESHOLD,
        "total_segments":     len(results),
        "qa_report_rows":     error_cnt,       # real errors only
        "exclusions_rows":    excl_cnt,        # everything else
        "low_conf_fp":        low_conf_cnt,    # subset of exclusions (yellow)
        "clean_count":        clean_cnt,       # subset of exclusions (green)
        "total_words":        total_words,
        "estimated_cost_usd": round(total_words * WORD_COST_USD, 4),
        "segments":           _results_to_json_list(results),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  XLIFF Parser
# ═══════════════════════════════════════════════════════════════════════════════

def _strip_ns(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def _text_of(elem: etree._Element) -> str:
    return (etree.tostring(elem, method="text", encoding="unicode") or "").strip()


class XliffParser:

    def parse(self, path: Path) -> List[XliffSegment]:
        try:
            tree = etree.parse(str(path))
            root = tree.getroot()
        except etree.XMLSyntaxError as exc:
            raise ValueError(f"XML parse error in {path.name}: {exc}") from exc

        root_local = _strip_ns(root.tag)
        if root_local == "xliff":
            return self._parse_xliff(root, path.name)
        raise ValueError(f"Unrecognised root element <{root_local}> in {path.name}")

    def _parse_xliff(
        self, root: etree._Element, file_name: str
    ) -> List[XliffSegment]:
        segments:    List[XliffSegment] = []
        seq_counter: int                = 0

        for file_elem in root.iter():
            if _strip_ns(file_elem.tag) != "file":
                continue
            doc_id = file_elem.get("original", file_name)

            for tu in file_elem.iter():
                if _strip_ns(tu.tag) != "trans-unit":
                    continue

                tu_id       = tu.get("id", "")
                source_text = ""
                target_text = ""
                state       = ""

                for child in tu:
                    local = _strip_ns(child.tag)
                    if local == "source":
                        source_text = _text_of(child)
                    elif local == "target":
                        target_text = _text_of(child)
                        state       = child.get("state", "")

                sdl_segs = self._extract_sdl_segments(tu)
                if sdl_segs:
                    for seg in sdl_segs:
                        seq_counter += 1
                        segments.append(XliffSegment(
                            seg_id    = seg.seg_id,
                            source    = seg.source    or source_text,
                            target    = seg.target    or target_text,
                            state     = seg.state     or state,
                            file_name = file_name,
                            doc_id    = doc_id,
                            seg_seq   = str(seq_counter),
                            extra     = seg.extra,
                        ))
                else:
                    if source_text or target_text:
                        seq_counter += 1
                        segments.append(XliffSegment(
                            seg_id    = tu_id,
                            source    = source_text,
                            target    = target_text,
                            state     = state,
                            file_name = file_name,
                            doc_id    = doc_id,
                            seg_seq   = str(seq_counter),
                        ))

        return segments

    def _extract_sdl_segments(
        self, tu: etree._Element
    ) -> List[XliffSegment]:
        seg_sources: Dict[str, str] = {}
        seg_targets: Dict[str, str] = {}
        seg_states:  Dict[str, str] = {}

        for child in tu:
            local = _strip_ns(child.tag)
            if local == "seg-source":
                for mrk in child.iter():
                    if _strip_ns(mrk.tag) == "mrk" and mrk.get("mtype") == "seg":
                        mid              = mrk.get("mid", "")
                        seg_sources[mid] = _text_of(mrk)
            elif local == "target":
                for mrk in child.iter():
                    if _strip_ns(mrk.tag) == "mrk" and mrk.get("mtype") == "seg":
                        mid   = mrk.get("mid", "")
                        state = (
                            mrk.get("{http://sdl.com/FileTypes/SdlXliff/1.0}state", "")
                            or mrk.get("state", "")
                        )
                        seg_targets[mid] = _text_of(mrk)
                        if state:
                            seg_states[mid] = state

        all_mids: List[str] = sorted(set(list(seg_sources) + list(seg_targets)))
        segments: List[XliffSegment] = []
        for mid in all_mids:
            tu_id  = tu.get("id", "")
            seg_id = f"{tu_id}::{mid}" if mid else tu_id
            segments.append(XliffSegment(
                seg_id    = seg_id,
                source    = seg_sources.get(mid, ""),
                target    = seg_targets.get(mid, ""),
                state     = seg_states.get(mid, ""),
                file_name = "",
                doc_id    = "",
                seg_seq   = "",
                extra     = {"tu_id": tu_id, "mrk_mid": mid},
            ))
        return segments


# ═══════════════════════════════════════════════════════════════════════════════
#  MosAIQ Report Reader  (context only)
# ═══════════════════════════════════════════════════════════════════════════════

class MosaiqReportReader:
    _COL_MAP: Dict[str, str] = {
        "doc-id":      "doc_id",
        "seg-seq":     "seg_seq",
        "seg-id":      "seg_id",
        "seg-info":    "seg_info",
        "source":      "source_text",
        "target":      "target_text",
        "fp":          "fp",
        "comment":     "comment",
        "issue sign":  "issue_sign",
        "issue group": "issue_group",
        "seg":         "seg_id",
        "seg id":      "seg_id",
        "segment id":  "seg_id",
        "id":          "seg_id",
        "source text": "source_text",
        "src":         "source_text",
        "target text": "target_text",
        "tgt":         "target_text",
        "translation": "target_text",
        "error":       "error_type",
        "error type":  "error_type",
        "issue":       "error_type",
        "issue type":  "error_type",
        "category":    "error_type",
        "check":       "error_type",
        "rule":        "error_type",
        "severity":    "severity",
        "level":       "severity",
        "weight":      "severity",
        "note":        "comment",
        "description": "comment",
        "message":     "comment",
        "detail":      "comment",
    }

    def read(self, path: Path) -> Tuple[List[MosaiqIssue], List[str], int]:
        try:
            wb   = load_workbook(str(path), read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            raise ValueError(f"Could not read MosAIQ report: {exc}") from exc

        if not rows:
            return [], [], 0

        header_idx = 0
        for i, row in enumerate(rows[:20]):
            non_empty = [c for c in row if c and isinstance(c, str)]
            if len(non_empty) >= 2:
                header_idx = i
                break

        header    = [str(c or "").strip() for c in rows[header_idx]]
        col_map:  Dict[int, str] = {}
        assigned: set            = set()

        for col_idx, col_name in enumerate(header):
            canonical = self._COL_MAP.get(col_name.lower())
            if canonical and canonical not in assigned:
                col_map[col_idx] = canonical
                assigned.add(canonical)

        issues: List[MosaiqIssue] = []
        for row_idx, row in enumerate(
            rows[header_idx + 1:], start=header_idx + 2
        ):
            if all(c is None or str(c).strip() == "" for c in row):
                continue

            mapped: Dict[str, str] = {}
            extra:  Dict[str, Any] = {}

            for col_idx, cell_val in enumerate(row):
                val       = str(cell_val).strip() if cell_val is not None else ""
                canonical = col_map.get(col_idx)
                if canonical:
                    mapped[canonical] = val
                else:
                    col_name = (
                        header[col_idx]
                        if col_idx < len(header)
                        else f"col_{col_idx}"
                    )
                    extra[col_name] = val

            issues.append(MosaiqIssue(
                row_index   = row_idx,
                seg_id      = mapped.get("seg_id",      ""),
                source_text = mapped.get("source_text", ""),
                target_text = mapped.get("target_text", ""),
                error_type  = mapped.get("error_type",  ""),
                severity    = mapped.get("severity",    ""),
                comment     = mapped.get("comment",     ""),
                doc_id      = mapped.get("doc_id",      ""),
                seg_seq     = mapped.get("seg_seq",     ""),
                seg_info    = mapped.get("seg_info",    ""),
                fp          = mapped.get("fp",          ""),
                issue_sign  = mapped.get("issue_sign",  ""),
                issue_group = mapped.get("issue_group", ""),
                extra       = extra,
            ))

        return issues, header, header_idx


# ═══════════════════════════════════════════════════════════════════════════════
#  Instructions parser
# ═══════════════════════════════════════════════════════════════════════════════

def parse_instructions_docx(path: Path) -> str:
    doc   = Document(str(path))
    lines = []
    for para in doc.paragraphs:
        t = para.text.strip()
        if t:
            lines.append(t)
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
#  AI evaluation
# ═══════════════════════════════════════════════════════════════════════════════

_EVAL_SYSTEM = """\
You are a senior translation quality-assurance specialist reviewing SDL XLIFF
translation segments.

Your job:
1. Examine the SOURCE and TARGET text of the XLIFF segment.
2. Using the client/project instructions AND the sample MosAIQ error patterns
   as guidance, decide whether the TARGET contains a genuine translation error.
3. Pay special attention to metric/imperial unit conversions (mm, cm, m, km,
   in, ft, yd, mi, degrees C, degrees F, kg, lb, oz, etc.).

Definitions
-----------
has_issue         : true  = the target text has a real translation problem.
is_false_positive : true  = an automated QA tool flagged this segment, but
                            the translation is actually correct.
                            Set ONLY when has_issue is false.
ai_confidence     : how confident YOU are in YOUR verdict (0.0 = not at all,
                    1.0 = certain).

Return ONLY valid JSON, no markdown, no extra text.
{
  "has_issue":         true | false,
  "is_false_positive": true | false,
  "ai_confidence":     0.0 to 1.0,
  "ai_issue":          "Brief label e.g. Incorrect unit conversion or None",
  "ai_comment":        "One-sentence QA warning or clearance note.",
  "ai_rationale":      "Two-to-four sentences explaining your decision."
}
"""


def _summarise_mosaiq_context(
    issues: List[MosaiqIssue], max_examples: int = 8
) -> str:
    if not issues:
        return "No MosAIQ reference data available."

    lines = ["MosAIQ QA error-pattern reference (sample):"]
    seen_types: Dict[str, int] = {}
    for iss in issues:
        et = iss.error_type or iss.issue_group or "Unknown"
        seen_types[et] = seen_types.get(et, 0) + 1

    lines.append("Error-type frequency in this report:")
    for et, cnt in sorted(seen_types.items(), key=lambda x: -x[1])[:10]:
        lines.append(f"  - {et}: {cnt} occurrence(s)")

    lines.append("\nExample flagged segments (for pattern reference only):")
    shown = 0
    for iss in issues:
        if shown >= max_examples:
            break
        if not iss.source_text and not iss.target_text:
            continue
        lines.append(
            f"  [{iss.error_type or iss.issue_group}] "
            f"SRC: {iss.source_text[:120]} | "
            f"TGT: {iss.target_text[:120]} | "
            f"Comment: {iss.comment[:80]}"
        )
        shown += 1

    return "\n".join(lines)


def _build_segment_prompt(
    segment:        XliffSegment,
    instructions:   str,
    mosaiq_context: str,
) -> str:
    return (
        f"PROJECT / CLIENT INSTRUCTIONS:\n"
        f"{instructions[:5000]}\n\n"
        f"{mosaiq_context}\n\n"
        f"XLIFF SEGMENT TO EVALUATE:\n"
        f"  File    : {segment.file_name}\n"
        f"  Doc-ID  : {segment.doc_id}\n"
        f"  Seg-Seq : {segment.seg_seq}\n"
        f"  Seg-ID  : {segment.seg_id}\n"
        f"  State   : {segment.state}\n"
        f"  Source  : {segment.source}\n"
        f"  Target  : {segment.target}\n\n"
        f"Evaluate the TARGET translation and return JSON as specified."
    )


def evaluate_single_segment(
    segment:        XliffSegment,
    instructions:   str,
    mosaiq_context: str,
    matched_mosaiq: Optional[MosaiqIssue],
    api_key:        str,
    model:          str,
    temperature:    float,
) -> EvaluatedSegment:
    prompt = _build_segment_prompt(segment, instructions, mosaiq_context)
    client = Anthropic(api_key=api_key)
    try:
        msg = client.messages.create(
            model=model,
            max_tokens=512,
            temperature=temperature,
            system=_EVAL_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        text_blocks = [
            getattr(item, "text", "")
            for item in msg.content
            if getattr(item, "type", "") == "text"
        ]
        raw  = "\n".join(text_blocks)
        data = safe_json_loads(raw) or {}

        return EvaluatedSegment(
            segment           = segment,
            has_issue         = bool(data.get("has_issue",         False)),
            is_false_positive = bool(data.get("is_false_positive", False)),
            ai_confidence     = float(data.get("ai_confidence",    0.5)),
            ai_issue          = str(data.get("ai_issue",           "None")),
            ai_comment        = str(data.get("ai_comment",         "No comment.")),
            ai_rationale      = str(data.get("ai_rationale",       "No rationale.")),
            matched_mosaiq    = matched_mosaiq,
        )
    except Exception as exc:
        return EvaluatedSegment(
            segment           = segment,
            has_issue         = False,
            is_false_positive = False,
            ai_confidence     = 0.0,
            ai_issue          = "Evaluation error",
            ai_comment        = f"Evaluation failed: {exc}",
            ai_rationale      = f"API call failed: {exc}",
            matched_mosaiq    = matched_mosaiq,
        )


def evaluate_segments_parallel(
    segments:       List[XliffSegment],
    instructions:   str,
    mosaiq_context: str,
    seg_to_mosaiq:  Dict[str, MosaiqIssue],
    api_key:        str,
    model:          str,
    temperature:    float,
    max_workers:    int = MAX_WORKERS,
    progress_cb     = None,
) -> List[EvaluatedSegment]:
    results: List[Optional[EvaluatedSegment]] = [None] * len(segments)
    total   = len(segments)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_idx = {
            executor.submit(
                evaluate_single_segment,
                seg,
                instructions,
                mosaiq_context,
                seg_to_mosaiq.get(seg.seg_id),
                api_key,
                model,
                temperature,
            ): idx
            for idx, seg in enumerate(segments)
        }

        done = 0
        for future in concurrent.futures.as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                results[idx] = future.result()
            except Exception as exc:
                results[idx] = EvaluatedSegment(
                    segment           = segments[idx],
                    has_issue         = False,
                    is_false_positive = False,
                    ai_confidence     = 0.0,
                    ai_issue          = "Worker error",
                    ai_comment        = f"Worker error: {exc}",
                    ai_rationale      = f"Worker error: {exc}",
                )
            done += 1
            if progress_cb:
                progress_cb(done, total)

    return [r for r in results if r is not None]


def build_seg_to_mosaiq(
    segments: List[XliffSegment],
    issues:   List[MosaiqIssue],
) -> Dict[str, MosaiqIssue]:
    by_id:  Dict[str, MosaiqIssue] = {
        iss.seg_id: iss for iss in issues if iss.seg_id
    }
    by_src: Dict[str, MosaiqIssue] = {}
    for iss in issues:
        key = iss.source_text[:80].strip().lower()
        if key:
            by_src[key] = iss

    result: Dict[str, MosaiqIssue] = {}
    for seg in segments:
        if seg.seg_id in by_id:
            result[seg.seg_id] = by_id[seg.seg_id]
        else:
            key = seg.source[:80].strip().lower()
            if key in by_src:
                result[seg.seg_id] = by_src[key]
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel Report Writer
# ═══════════════════════════════════════════════════════════════════════════════

_HDR_FILL     = PatternFill("solid", fgColor="1F6B3A")
_HDR_FONT     = Font(bold=True, color="FFFFFF", name="Segoe UI", size=9)
_WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
_REAL_FILL    = PatternFill("solid", fgColor="FFCCCC")
_FP_FILL      = PatternFill("solid", fgColor="FFFFFF")
_LOWCONF_FILL = PatternFill("solid", fgColor="FFF3CC")
_CLEAN_FILL   = PatternFill("solid", fgColor="E8F5E9")
_PLAIN_FILL   = PatternFill("solid", fgColor="FFFFFF")
_BODY_FONT    = Font(name="Segoe UI", size=9, color="1e1f2e")

_SPEC_HEADERS = [
    "Doc-ID", "Seg-Seq", "Seg-ID", "Seg-Info",
    "Source", "Target",
    "FP", "Comment", "Issue Sign", "Issue Group",
]
_AI_HEADERS = [
    "AI Issue",       # K
    "AI Confidence",  # L  — numeric 0.000–1.000
    "AI Comment",     # M
    "AI Rationale",   # N
]
_ALL_HEADERS = _SPEC_HEADERS + _AI_HEADERS


def _style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell           = ws.cell(row=1, column=col)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = _WRAP_ALIGN


def _auto_col_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length     = max(len(str(c.value or "")) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 2, max_w))


def _append_row(ws, values: list, fill: PatternFill) -> None:
    ws.append(values)
    row_idx = ws.max_row
    for col_idx in range(1, len(values) + 1):
        cell           = ws.cell(row=row_idx, column=col_idx)
        cell.fill      = fill
        cell.font      = _BODY_FONT
        cell.alignment = _WRAP_ALIGN


def _build_row(result: EvaluatedSegment) -> list:
    seg = result.segment
    mq  = result.matched_mosaiq

    if result.is_false_positive:
        fp_val = "Yes"
    elif result.has_issue:
        fp_val = "No"
    else:
        fp_val = ""

    ai_comment = result.ai_comment
    if result.is_false_positive and result.ai_confidence < FP_CONFIDENCE_THRESHOLD:
        ai_comment = (
            f"[LOW CONFIDENCE – HUMAN REVIEW REQUIRED "
            f"(score={result.ai_confidence:.2f})] {ai_comment}"
        )

    return [
        seg.doc_id,
        seg.seg_seq,
        seg.seg_id,
        seg.state,
        seg.source,
        seg.target,
        fp_val,
        mq.comment     if mq else "",
        mq.issue_sign  if mq else "",
        mq.issue_group if mq else "",
        result.ai_issue,
        round(result.ai_confidence, 3),
        ai_comment,
        result.ai_rationale,
    ]


def _pick_fill(result: EvaluatedSegment) -> PatternFill:
    if result.has_issue:
        return _REAL_FILL
    if result.is_false_positive:
        return _FP_FILL if result.ai_confidence >= FP_CONFIDENCE_THRESHOLD else _LOWCONF_FILL
    return _CLEAN_FILL


def write_metric_report(
    out_path:    Path,
    results:     List[EvaluatedSegment],
    total_words: int = 0,
) -> Tuple[int, int, int]:
    """
    Sheet layout:
      QA_Report  — real errors only          (red rows)
      Exclusions — all non-errors:
                     high-conf FP → white
                     low-conf FP  → yellow
                     clean segs   → green
      All_Issues — every segment
      Cost       — word count / cost estimate

    QA_Report count + Exclusions count == len(results)  always.
    """
    wb = Workbook()

    # ── classify ──────────────────────────────────────────────────────────────
    qa_rows:   List[EvaluatedSegment] = []   # real errors
    excl_rows: List[EvaluatedSegment] = []   # everything that is NOT a real error

    for r in results:
        if r.has_issue:
            qa_rows.append(r)
        else:
            excl_rows.append(r)   # confirmed FP, low-conf FP, and clean all go here

    # ── Sheet 1 : QA_Report  (real errors only) ───────────────────────────────
    ws_qa       = wb.active
    ws_qa.title = "QA_Report"
    ws_qa.append(_ALL_HEADERS)
    _style_header_row(ws_qa, len(_ALL_HEADERS))
    for r in qa_rows:
        _append_row(ws_qa, _build_row(r), _REAL_FILL)
    _auto_col_width(ws_qa)

    # ── Sheet 2 : Exclusions  (all non-errors) ────────────────────────────────
    ws_excl = wb.create_sheet("Exclusions")
    ws_excl.append(_ALL_HEADERS)
    _style_header_row(ws_excl, len(_ALL_HEADERS))
    for r in excl_rows:
        # high-confidence FP  → white
        # low-confidence FP   → yellow  (needs human review)
        # clean (no issue)    → green
        if r.is_false_positive and r.ai_confidence < FP_CONFIDENCE_THRESHOLD:
            fill = _LOWCONF_FILL   # yellow
        elif r.is_false_positive:
            fill = _FP_FILL        # white
        else:
            fill = _CLEAN_FILL     # green
        _append_row(ws_excl, _build_row(r), fill)
    _auto_col_width(ws_excl)

    # ── Sheet 3 : All_Issues  (every segment, same colours) ───────────────────
    ws_all = wb.create_sheet("All_Issues")
    ws_all.append(_ALL_HEADERS)
    _style_header_row(ws_all, len(_ALL_HEADERS))
    for r in results:
        _append_row(ws_all, _build_row(r), _pick_fill(r))
    _auto_col_width(ws_all)

    # ── Sheet 4 : Cost ────────────────────────────────────────────────────────
    ws_cost = wb.create_sheet("Cost")
    ws_cost.append(["Total Words", "Rate USD/Word", "Estimated Cost USD"])
    _style_header_row(ws_cost, 3)
    _append_row(
        ws_cost,
        [total_words, WORD_COST_USD, round(total_words * WORD_COST_USD, 4)],
        _PLAIN_FILL,
    )
    _auto_col_width(ws_cost)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    return len(qa_rows), len(excl_rows), 0   # third value kept for signature compat


# ═══════════════════════════════════════════════════════════════════════════════
#  Custom Tkinter Widgets
# ═══════════════════════════════════════════════════════════════════════════════

class FlatButton:
    def __init__(
        self, parent, text="", command=None,
        bg=ACCENT, hover_bg=ACCENT_HOVER,
        fg=TEXT_PRIMARY, font=("Segoe UI", 9, "bold"),
        width=130, height=32, radius=8,
    ):
        try:
            parent_bg = parent.cget("bg")
        except Exception:
            parent_bg = BG_DARK

        self._text    = text
        self._command = command
        self._bg      = bg
        self._hbg     = hover_bg
        self._fg      = fg
        self._font    = font
        self._radius  = radius
        self._w       = width
        self._h       = height

        self.widget = tk.Canvas(
            parent, width=width, height=height,
            bg=parent_bg, highlightthickness=0, bd=0, cursor="hand2",
        )
        self.widget.after(0, lambda: self._draw(self._bg))
        self.widget.bind("<Enter>",    lambda _: self._draw(self._hbg))
        self.widget.bind("<Leave>",    lambda _: self._draw(self._bg))
        self.widget.bind("<Button-1>", lambda _: self._on_click())

    def pack(self, **kw):   self.widget.pack(**kw)
    def grid(self, **kw):   self.widget.grid(**kw)
    def place(self, **kw):  self.widget.place(**kw)
    def config(self, **kw): self.widget.config(**kw)

    def _round_rect(self, x1, y1, x2, y2, r, **kw):
        pts = [
            x1+r, y1,   x2-r, y1,   x2,   y1,
            x2,   y1+r, x2,   y2-r, x2,   y2,
            x2-r, y2,   x1+r, y2,   x1,   y2,
            x1,   y2-r, x1,   y1+r, x1,   y1,
            x1+r, y1,
        ]
        return self.widget.create_polygon(pts, smooth=True, **kw)

    def _draw(self, color: str) -> None:
        self.widget.delete("all")
        self._round_rect(0, 0, self._w, self._h, self._radius, fill=color, outline="")
        self.widget.create_text(
            self._w // 2, self._h // 2,
            text=self._text, fill=self._fg,
            font=self._font, anchor="center",
        )

    def _on_click(self) -> None:
        self._draw(self._bg)
        if self._command:
            self._command()


class PathRow(tk.Frame):
    _STATE_FILE:   Path = Path.home() / ".qa_reviewer_dirs.json"
    _state:        dict = {}
    _state_loaded: bool = False

    @classmethod
    def _load_state(cls) -> None:
        if cls._state_loaded:
            return
        cls._state_loaded = True
        try:
            if cls._STATE_FILE.exists():
                raw        = json.loads(cls._STATE_FILE.read_text(encoding="utf-8"))
                cls._state = {k: v for k, v in raw.items() if Path(v).exists()}
        except Exception:
            cls._state = {}

    @classmethod
    def _save_state(cls) -> None:
        try:
            cls._STATE_FILE.write_text(
                json.dumps(cls._state, indent=2, ensure_ascii=False),
                encoding="utf-8",
            )
        except Exception:
            pass

    @classmethod
    def _get_global_last(cls) -> str:
        return cls._state.get("__last_global__", "")

    @classmethod
    def _set_global_last(cls, directory: str) -> None:
        cls._state["__last_global__"] = directory
        cls._save_state()

    def __init__(
        self, parent, label: str, icon: str,
        pick_mode:  str = "folder",
        file_types      = None,
        row_key:    str = "",
        **kwargs,
    ):
        super().__init__(parent, bg=BG_CARD, **kwargs)
        PathRow._load_state()

        self._mode      = pick_mode
        self._filetypes = file_types or [("All", "*.*")]
        self._row_key   = row_key or label
        self._initdir   = self._resolve_initdir()

        lbl_frame = tk.Frame(self, bg=BG_CARD)
        lbl_frame.pack(fill=tk.X, padx=12, pady=(10, 2))
        tk.Label(lbl_frame, text=icon, bg=BG_CARD,
                 fg=ACCENT, font=("Segoe UI", 11)).pack(side=tk.LEFT)
        tk.Label(lbl_frame, text=f"  {label}", bg=BG_CARD,
                 fg=TEXT_SECONDARY, font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT)

        row = tk.Frame(self, bg=BG_CARD)
        row.pack(fill=tk.X, padx=12, pady=(0, 10))

        entry_wrap = tk.Frame(row, bg=BORDER, bd=0)
        entry_wrap.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        inner = tk.Frame(entry_wrap, bg=BG_INPUT, bd=0)
        inner.pack(fill=tk.BOTH, padx=1, pady=1)

        self.var    = tk.StringVar()
        self._entry = tk.Entry(
            inner, textvariable=self.var,
            bg=BG_INPUT, fg=TEXT_PRIMARY,
            insertbackground=TEXT_PRIMARY,
            relief=tk.FLAT, bd=6,
            font=("Segoe UI", 9),
            disabledbackground=BG_INPUT,
            disabledforeground=TEXT_SECONDARY,
        )
        self._entry.pack(fill=tk.X)
        self._entry.bind("<FocusIn>",  lambda _: entry_wrap.config(bg=ACCENT))
        self._entry.bind("<FocusOut>", lambda _: entry_wrap.config(bg=BORDER))
        self.var.trace_add("write", self._on_entry_change)

        FlatButton(
            row, text="Browse",
            bg=ACCENT_DIM, hover_bg=ACCENT,
            width=90, height=30, command=self._pick,
        ).pack(side=tk.LEFT)

    def _resolve_initdir(self) -> str:
        row_saved = PathRow._state.get(self._row_key, "")
        if row_saved and Path(row_saved).exists():
            return row_saved
        global_last = self._get_global_last()
        if global_last and Path(global_last).exists():
            return global_last
        desktop = Path.home() / "Desktop"
        if desktop.exists():
            return str(desktop)
        for candidate in [
            Path.home() / "Documents",
            *sorted(Path.home().glob("OneDrive*/Documents")),
        ]:
            if candidate.exists():
                return str(candidate)
        return str(Path.home())

    def _remember(self, directory: str) -> None:
        if not directory or not Path(directory).exists():
            return
        PathRow._state[self._row_key] = directory
        self._set_global_last(directory)

    def _on_entry_change(self, *_) -> None:
        typed = self.var.get().strip()
        if not typed:
            return
        p         = Path(typed)
        candidate = p.parent if (p.is_file() or (not p.exists() and p.suffix)) else p
        if candidate.exists():
            self._initdir = str(candidate)

    def get(self) -> str:
        return self.var.get().strip()

    def set(self, value: str) -> None:
        self.var.set(value)
        if value:
            p             = Path(value)
            new_dir       = str(p.parent if p.is_file() else p)
            self._initdir = new_dir
            self._remember(new_dir)

    def _pick(self) -> None:
        current = self.var.get().strip()
        if current:
            p         = Path(current)
            candidate = p.parent if p.is_file() else p
            if candidate.exists():
                self._initdir = str(candidate)

        if self._mode == "folder":
            selected = filedialog.askdirectory(
                initialdir=self._initdir, title="Select folder"
            )
        else:
            selected = filedialog.askopenfilename(
                initialdir=self._initdir,
                filetypes=self._filetypes,
                title="Select file",
            )
        if selected:
            self.set(selected)


class StatusBar(tk.Frame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, bg=BG_DARK, height=28, **kwargs)
        self.pack_propagate(False)
        self._dot = tk.Label(self, text="●", bg=BG_DARK,
                             fg=TEXT_MUTED, font=("Segoe UI", 10))
        self._dot.pack(side=tk.LEFT, padx=(12, 4))
        self._lbl = tk.Label(self, text="Ready", bg=BG_DARK,
                             fg=TEXT_SECONDARY, font=("Segoe UI", 8))
        self._lbl.pack(side=tk.LEFT)
        self._time_lbl = tk.Label(self, text="", bg=BG_DARK,
                                  fg=TEXT_MUTED, font=("Segoe UI", 8))
        self._time_lbl.pack(side=tk.RIGHT, padx=12)

    def set(self, text: str, state: str = "idle") -> None:
        colours = {
            "idle":    TEXT_MUTED,
            "running": WARNING,
            "ok":      SUCCESS,
            "error":   DANGER,
        }
        self._dot.config(fg=colours.get(state, TEXT_MUTED))
        self._lbl.config(
            text=text,
            fg=TEXT_PRIMARY if state != "idle" else TEXT_SECONDARY,
        )
        self._time_lbl.config(text=datetime.now().strftime("%H:%M:%S"))


# ═══════════════════════════════════════════════════════════════════════════════
#  Tab 1 — QA Reviewer shell (disabled)
# ═══════════════════════════════════════════════════════════════════════════════

class Tab1Frame(tk.Frame):
    def __init__(self, parent: tk.Widget, status_bar: StatusBar, **kwargs) -> None:
        super().__init__(parent, bg=BG_DARK, **kwargs)
        self._status_bar = status_bar
        self._build_ui()

    def _build_ui(self) -> None:
        body = tk.Frame(self, bg=BG_DARK)
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        left = tk.Frame(body, bg=BG_DARK, width=440)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0, 8))
        left.pack_propagate(False)

        right = tk.Frame(body, bg=BG_DARK)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        def section_label(p: tk.Frame, text: str) -> None:
            f = tk.Frame(p, bg=BG_DARK)
            f.pack(fill=tk.X, pady=(10, 4))
            tk.Label(f, text=text, bg=BG_DARK, fg=ACCENT,
                     font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT)
            tk.Frame(f, bg=BORDER, height=1).pack(
                side=tk.LEFT, fill=tk.X, expand=True, padx=8, pady=6)

        section_label(left, "INPUT FOLDERS")
        PathRow(left, "Article Folder (.docx)",             "📄", row_key="t1_articles").pack(fill=tk.X, pady=2)
        PathRow(left, "Reference Materials Folder (.docx)", "📚", row_key="t1_references").pack(fill=tk.X, pady=2)
        PathRow(left, "Internal Spot Check Folder (.docx)", "🔍", row_key="t1_spotchecks").pack(fill=tk.X, pady=2)

        section_label(left, "OUTPUT & CONFIGURATION")
        PathRow(left, "Output Folder", "📁", row_key="t1_output").pack(fill=tk.X, pady=2)
        PathRow(left, ".env File  (ANTHROPIC_API_KEY)", "🔑",
                pick_mode="file",
                file_types=[("ENV file", ".env"), ("All files", "*.*")],
                row_key="t1_env").pack(fill=tk.X, pady=2)

        btn_frame = tk.Frame(left, bg=BG_DARK)
        btn_frame.pack(fill=tk.X, pady=(18, 4))
        FlatButton(btn_frame, text="▶  Run QA Workflow",
                   command=self._disabled_notice,
                   bg=ACCENT_DIM, hover_bg=ACCENT_DIM,
                   width=200, height=38,
                   font=("Segoe UI", 10, "bold")).pack(side=tk.RIGHT)

        section_label(left, "AGENT PIPELINE")
        self._build_agent_cards(left)

        log_header = tk.Frame(right, bg=BG_PANEL)
        log_header.pack(fill=tk.X)
        tk.Label(log_header, text="  ◈  Activity Log", bg=BG_PANEL,
                 fg=TEXT_PRIMARY, font=("Segoe UI", 9, "bold")).pack(
            side=tk.LEFT, pady=8, padx=4)
        tk.Frame(right, bg=BORDER, height=1).pack(fill=tk.X)

        log_wrap = tk.Frame(right, bg=BG_INPUT)
        log_wrap.pack(fill=tk.BOTH, expand=True)
        self.log_box = tk.Text(log_wrap, bg=BG_INPUT, fg=TEXT_SECONDARY,
                               font=("Consolas", 9), relief=tk.FLAT, bd=10,
                               wrap=tk.WORD, state=tk.DISABLED)
        self.log_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb = ttk.Scrollbar(log_wrap, orient=tk.VERTICAL, command=self.log_box.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_box.config(yscrollcommand=sb.set)
        self.log_box.config(state=tk.NORMAL)
        self.log_box.insert(
            tk.END,
            "Tab 1 — QA Reviewer\n\n"
            "This tab's workflow is currently disabled.\n"
            "Please use Tab 2 (METRIC SYSTEM) for active processing.\n",
        )
        self.log_box.config(state=tk.DISABLED)

    def _disabled_notice(self) -> None:
        messagebox.showinfo(
            "Disabled",
            "Tab 1 workflow is disabled in this build.\n"
            "Please use the METRIC SYSTEM tab.",
        )

    def _build_agent_cards(self, parent: tk.Frame) -> None:
        cards_frame = tk.Frame(parent, bg=BG_DARK)
        cards_frame.pack(fill=tk.X, pady=2)
        for title, model, color, icon, side_pad in [
            ("Agent 1", "Claude Opus — Rules", "#c97bf7", "🧠", (0, 4)),
            ("Agent 2", "Claude Haiku — QA",   "#c97bf7", "🧠", (4, 0)),
        ]:
            card  = tk.Frame(cards_frame, bg=BG_CARD)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=side_pad)
            strip = tk.Frame(card, bg=color, width=3)
            strip.pack(side=tk.LEFT, fill=tk.Y)
            strip.pack_propagate(False)
            inner = tk.Frame(card, bg=BG_CARD, padx=10, pady=8)
            inner.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            tk.Label(inner, text=f"{icon} {title}", bg=BG_CARD,
                     fg=color, font=("Segoe UI", 8, "bold")).pack(anchor="w")
            tk.Label(inner, text=model, bg=BG_CARD,
                     fg=TEXT_SECONDARY, font=("Segoe UI", 8)).pack(anchor="w")


# ═══════════════════════════════════════════════════════════════════════════════
#  Tab 2 — METRIC SYSTEM
# ═══════════════════════════════════════════════════════════════════════════════

class Tab2Frame(tk.Frame):
    def __init__(self, parent: tk.Widget, status_bar: StatusBar, **kwargs) -> None:
        super().__init__(parent, bg=BG_DARK, **kwargs)
        self._status_bar  = status_bar
        self._total_files = 1
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        body = tk.Frame(self, bg=BG_DARK)
        body.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        left = tk.Frame(body, bg=BG_DARK, width=460)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(0, 8))
        left.pack_propagate(False)

        right = tk.Frame(body, bg=BG_DARK)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        def section_label(p: tk.Frame, text: str) -> None:
            f = tk.Frame(p, bg=BG_DARK)
            f.pack(fill=tk.X, pady=(10, 4))
            tk.Label(f, text=text, bg=BG_DARK, fg=ACCENT,
                     font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT)
            tk.Frame(f, bg=BORDER, height=1).pack(
                side=tk.LEFT, fill=tk.X, expand=True, padx=8, pady=6)

        section_label(left, "INPUT FILES")

        self.row_xliff = PathRow(
            left, "XLIFF File  (SDL XLIFF / XLIFF 1.2 / MXLIFF)", "🌐",
            pick_mode="file",
            file_types=[
                ("XLIFF files", "*.sdlxliff *.xliff *.mxliff *.xlf"),
                ("All files",   "*.*"),
            ],
            row_key="t2_xliff",
        )
        self.row_xliff.pack(fill=tk.X, pady=2)

        self.row_xliff_folder = PathRow(
            left, "— OR — SDL XLIFF Folder  (batch, *.sdlxliff)", "📂",
            pick_mode="folder",
            row_key="t2_xliff_folder",
        )
        self.row_xliff_folder.pack(fill=tk.X, pady=2)

        tk.Label(
            left,
            text="  ⓘ  Single file takes priority.  Folder mode → one Excel per SDL XLIFF file.",
            bg=BG_DARK, fg=TEXT_MUTED, font=("Segoe UI", 7),
        ).pack(anchor="w", padx=14, pady=(0, 4))

        self.row_instructions = PathRow(
            left, "Client / Project Instructions  (.docx)", "📋",
            pick_mode="file",
            file_types=[("Word documents", "*.docx"), ("All files", "*.*")],
            row_key="t2_instructions",
        )
        self.row_instructions.pack(fill=tk.X, pady=2)

        self.row_report = PathRow(
            left, "MosAIQ QA Report  (.xlsx)  — context only", "📊",
            pick_mode="file",
            file_types=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            row_key="t2_report",
        )
        self.row_report.pack(fill=tk.X, pady=2)

        section_label(left, "OUTPUT & CONFIGURATION")

        self.row_output = PathRow(left, "Output Folder", "📁", row_key="t2_output")
        self.row_output.pack(fill=tk.X, pady=2)
        self.row_output.set(str(Path.cwd() / "output_metric"))

        self.row_env = PathRow(
            left, ".env File  (ANTHROPIC_API_KEY)", "🔑",
            pick_mode="file",
            file_types=[("ENV file", ".env"), ("All files", "*.*")],
            row_key="t2_env",
        )
        self.row_env.pack(fill=tk.X, pady=2)

        # colour legend
        legend_frame = tk.Frame(left, bg=BG_DARK)
        legend_frame.pack(fill=tk.X, pady=(6, 0))
        tk.Label(legend_frame, text="Row colours: ", bg=BG_DARK,
                 fg=TEXT_MUTED, font=("Segoe UI", 7)).pack(side=tk.LEFT)
        for colour, label in [
            ("#FFCCCC", "Real error"),
            ("#FFF3CC", "Low-conf FP"),
            ("#FFFFFF", "Confirmed FP"),
            ("#E8F5E9", "Clean"),
        ]:
            tk.Label(legend_frame, text="■", bg=BG_DARK,
                     fg=colour, font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(6, 0))
            tk.Label(legend_frame, text=label, bg=BG_DARK,
                     fg=TEXT_MUTED, font=("Segoe UI", 7)).pack(side=tk.LEFT)

        # run button
        btn_frame = tk.Frame(left, bg=BG_DARK)
        btn_frame.pack(fill=tk.X, pady=(14, 4))
        self._run_btn = FlatButton(
            btn_frame, text="▶  Run Metric Filter",
            command=self._run,
            bg=ACCENT, hover_bg=ACCENT_HOVER,
            width=210, height=38,
            font=("Segoe UI", 10, "bold"),
        )
        self._run_btn.pack(side=tk.RIGHT)

        section_label(left, "AGENT PIPELINE")
        self._build_agent_cards(left)

        # stats strip
        self._stats_frame = tk.Frame(left, bg=BG_CARD)
        self._stats_frame.pack(fill=tk.X, pady=(8, 0))
        self._lbl_total = self._stat_label(self._stats_frame, "Segments", "—")
        self._lbl_real  = self._stat_label(self._stats_frame, "Errors",   "—")
        self._lbl_excl  = self._stat_label(self._stats_frame, "FP Excl.", "—")
        self._lbl_clean = self._stat_label(self._stats_frame, "Clean",    "—")

        # progress bar
        prog_frame = tk.Frame(left, bg=BG_DARK)
        prog_frame.pack(fill=tk.X, pady=(4, 0))
        style = ttk.Style()
        style.configure(
            "M2.Horizontal.TProgressbar",
            troughcolor=BG_PANEL, background=ACCENT,
            bordercolor=BG_PANEL, lightcolor=ACCENT, darkcolor=ACCENT,
        )
        self._progress = ttk.Progressbar(
            prog_frame, style="M2.Horizontal.TProgressbar",
            orient=tk.HORIZONTAL, mode="determinate", length=200,
        )
        self._progress.pack(fill=tk.X, pady=2)
        self._progress_lbl = tk.Label(prog_frame, text="", bg=BG_DARK,
                                       fg=TEXT_MUTED, font=("Segoe UI", 7))
        self._progress_lbl.pack(anchor="e", padx=2)

        # log panel
        log_header = tk.Frame(right, bg=BG_PANEL)
        log_header.pack(fill=tk.X)
        tk.Label(log_header, text="  ◈  Activity Log", bg=BG_PANEL,
                 fg=TEXT_PRIMARY, font=("Segoe UI", 9, "bold")).pack(
            side=tk.LEFT, pady=8, padx=4)

        clear_btn = tk.Label(log_header, text="Clear  ✕", bg=BG_PANEL,
                             fg=TEXT_MUTED, font=("Segoe UI", 8), cursor="hand2")
        clear_btn.pack(side=tk.RIGHT, padx=10)
        clear_btn.bind("<Button-1>", lambda _: self._clear_log())
        clear_btn.bind("<Enter>",    lambda _: clear_btn.config(fg=TEXT_PRIMARY))
        clear_btn.bind("<Leave>",    lambda _: clear_btn.config(fg=TEXT_MUTED))

        tk.Frame(right, bg=BORDER, height=1).pack(fill=tk.X)

        log_wrap = tk.Frame(right, bg=BG_INPUT)
        log_wrap.pack(fill=tk.BOTH, expand=True)
        self.log_box = tk.Text(
            log_wrap, bg=BG_INPUT, fg=TEXT_PRIMARY,
            insertbackground=TEXT_PRIMARY,
            font=("Consolas", 9), relief=tk.FLAT, bd=10,
            wrap=tk.WORD, state=tk.DISABLED,
            selectbackground=ACCENT_DIM,
        )
        self.log_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.log_box.tag_config("time",    foreground=TEXT_MUTED)
        self.log_box.tag_config("info",    foreground=TEXT_PRIMARY)
        self.log_box.tag_config("success", foreground=SUCCESS)
        self.log_box.tag_config("warning", foreground=WARNING)
        self.log_box.tag_config("error",   foreground=DANGER)
        self.log_box.tag_config("agent",   foreground="#c97bf7")
        self.log_box.tag_config("dim",     foreground=TEXT_MUTED)

        sb = ttk.Scrollbar(log_wrap, orient=tk.VERTICAL, command=self.log_box.yview)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_box.config(yscrollcommand=sb.set)

    # ── progress callback factory (never nested inside try) ───────────────────
    def _make_progress_cb(self, file_idx: int, total_files: int):
        file_offset = (file_idx - 1) / max(total_files, 1)
        file_share  = 1.0 / max(total_files, 1)

        def _cb(done: int, total: int) -> None:
            if total == 0:
                return
            pct = int((file_offset + file_share * done / total) * 100)
            self.after(
                0,
                lambda p=pct, d=done, t=total: self._set_progress_abs(p, d, t),
            )

        return _cb

    # ── core workflow ─────────────────────────────────────────────────────────
    def _run_inner(self) -> None:
        try:
            self._status_bar.set("Metric System — running…", "running")
            self._progress["value"] = 0
            self._progress_lbl.config(text="")

            xliff_single = self.row_xliff.get()
            xliff_folder = self.row_xliff_folder.get()
            inst_path    = Path(self.row_instructions.get())
            report_path  = Path(self.row_report.get())
            out_dir      = Path(self.row_output.get())
            env_file     = Path(self.row_env.get())

            # ── validate inputs ───────────────────────────────────────────────
            errors: List[str] = []
            if not xliff_single and not xliff_folder:
                errors.append(
                    "Provide either a single XLIFF file or "
                    "a folder of SDL XLIFF files."
                )
            if xliff_single and not Path(xliff_single).exists():
                errors.append(f"XLIFF file not found: {xliff_single}")
            if xliff_folder and not Path(xliff_folder).is_dir():
                errors.append(f"XLIFF folder not found: {xliff_folder}")
            if not inst_path.exists():
                errors.append("Instructions file not found.")
            if not report_path.exists():
                errors.append("MosAIQ report not found.")
            if not env_file.exists():
                errors.append(".env file not found.")
            if errors:
                raise ValueError("\n".join(errors))

            env           = load_env_file(env_file)
            anthropic_key = env.get("ANTHROPIC_API_KEY", "")
            if not anthropic_key:
                raise ValueError("ANTHROPIC_API_KEY not found in .env file.")

            agent_cfg   = load_agent_settings(DEFAULT_MOSAIQ_CONFIG)
            model       = agent_cfg.get("step2_model", DEFAULT_CLAUDE_MODEL_AGENT2)
            temperature = float(agent_cfg.get("step2_temp", 0.2))

            out_dir.mkdir(parents=True, exist_ok=True)
            json_dir = out_dir / "json_cache"
            json_dir.mkdir(parents=True, exist_ok=True)

            # ── collect XLIFF paths ───────────────────────────────────────────
            if xliff_single:
                xliff_paths = [Path(xliff_single)]
                self._log(f"Single XLIFF mode → {Path(xliff_single).name}")
            else:
                folder      = Path(xliff_folder)
                xliff_paths = sorted(folder.rglob("*.sdlxliff"))
                if not xliff_paths:
                    raise ValueError(f"No *.sdlxliff files found in:\n{folder}")
                self._log(
                    f"Folder mode → {len(xliff_paths)} SDL XLIFF file(s) "
                    f"in {folder.name}"
                )

            self._total_files = len(xliff_paths)

            # ── parse instructions (shared across all files) ──────────────────
            self._log(f"Parsing instructions → {inst_path.name}")
            instructions_text = parse_instructions_docx(inst_path)
            inst_words        = word_count(instructions_text)
            self._log(f"  {inst_words} words extracted")
            save_json(
                json_dir / "instructions.json",
                instructions_to_json(instructions_text, inst_path.name),
            )

            # ── parse MosAIQ report for context (shared across all files) ─────
            self._log(f"Reading MosAIQ context → {report_path.name}")
            reader = MosaiqReportReader()
            mosaiq_issues, _, header_row_idx = reader.read(report_path)
            self._log(
                f"  {len(mosaiq_issues)} MosAIQ reference row(s) "
                f"(headers at row {header_row_idx + 1}) — context only"
            )
            mosaiq_context = _summarise_mosaiq_context(mosaiq_issues)

            # ── grand totals across all files ─────────────────────────────────
            ts              = datetime.now().strftime("%Y%m%d_%H%M%S")
            grand_total     = 0
            grand_errors    = 0
            grand_excl      = 0
            grand_low_conf  = 0
            grand_clean     = 0

            # ══════════════════════════════════════════════════════════════════
            # Process every XLIFF file independently
            # ══════════════════════════════════════════════════════════════════
            for file_idx, xp in enumerate(xliff_paths, start=1):
                self._log(
                    f"─── File {file_idx}/{len(xliff_paths)}: {xp.name} ───",
                    "agent",
                )

                # ── parse XLIFF → segments ────────────────────────────────────
                parser   = XliffParser()
                segments = parser.parse(xp)

                if not segments:
                    self._log(
                        f"  No segments found — skipping {xp.name}", "warning"
                    )
                    continue

                file_words = sum(word_count(s.source) for s in segments)
                self._log(
                    f"  {len(segments)} segment(s), {file_words} source words"
                )
                save_json(
                    json_dir / f"{xp.stem}_segments.json",
                    segments_to_json(segments),
                )

                # ── cross-reference segments with MosAIQ rows (best-effort) ───
                seg_to_mosaiq = build_seg_to_mosaiq(segments, mosaiq_issues)
                self._log(
                    f"  Cross-referenced {len(seg_to_mosaiq)} / "
                    f"{len(segments)} segments with MosAIQ rows"
                )

                # ── AI evaluation ─────────────────────────────────────────────
                self._log(
                    f"  Evaluating {len(segments)} segment(s) via Claude "
                    f"[model={model}, workers={MAX_WORKERS}]…",
                    "agent",
                )

                results = evaluate_segments_parallel(
                    segments       = segments,
                    instructions   = instructions_text,
                    mosaiq_context = mosaiq_context,
                    seg_to_mosaiq  = seg_to_mosaiq,
                    api_key        = anthropic_key,
                    model          = model,
                    temperature    = temperature,
                    max_workers    = MAX_WORKERS,
                    progress_cb    = self._make_progress_cb(
                        file_idx, len(xliff_paths)
                    ),
                )

                # ── tally results ─────────────────────────────────────────────
                # QA_Report  = real errors only
                # Exclusions = everything else (confirmed FP + low-conf FP + clean)
                # QA + Excl  = total   (always)
                error_cnt    = sum(1 for r in results if r.has_issue)
                excl_cnt     = len(results) - error_cnt
                low_conf_cnt = sum(
                    1 for r in results
                    if r.is_false_positive
                    and r.ai_confidence < FP_CONFIDENCE_THRESHOLD
                )
                clean_cnt    = sum(
                    1 for r in results
                    if not r.has_issue and not r.is_false_positive
                )

                self._log(
                    f"  → QA_Report: {error_cnt}  |  "
                    f"Exclusions: {excl_cnt} "
                    f"({low_conf_cnt} low-conf FP yellow, "
                    f"{clean_cnt} clean green)  |  "
                    f"Total: {len(results)}",
                    "success",
                )

                # sanity check
                assert error_cnt + excl_cnt == len(results), (
                    f"QA_Report ({error_cnt}) + Exclusions ({excl_cnt}) "
                    f"!= Total ({len(results)})"
                )

                # ── write Excel named after the XLIFF stem ────────────────────
                total_words = file_words + inst_words
                write_metric_report(
                    out_path    = out_dir / f"{xp.stem}.xlsx",
                    results     = results,
                    total_words = total_words,
                )
                self._log(f"  Saved → {xp.stem}.xlsx", "success")

                # ── per-file JSON (uses standalone helpers — no open literals) ─
                file_json = _build_file_json(
                    ts           = ts,
                    xp           = xp,
                    inst_path    = inst_path,
                    report_path  = report_path,
                    model        = model,
                    results      = results,
                    error_cnt    = error_cnt,
                    excl_cnt     = excl_cnt,
                    low_conf_cnt = low_conf_cnt,
                    clean_cnt    = clean_cnt,
                    total_words  = total_words,
                )
                save_json(
                    out_dir / f"{xp.stem}_results_{ts}.json",
                    file_json,
                )

                # ── accumulate grand totals ────────────────────────────────────
                grand_total    += len(results)
                grand_errors   += error_cnt
                grand_excl     += excl_cnt
                grand_low_conf += low_conf_cnt
                grand_clean    += clean_cnt

            # ══════════════════════════════════════════════════════════════════
            # All files done
            # ══════════════════════════════════════════════════════════════════
            self.after(
                0,
                lambda: self._update_stats(
                    grand_total,
                    grand_errors,
                    grand_excl,
                    grand_clean,
                ),
            )
            self._set_progress_abs(100, grand_total, grand_total)
            self._status_bar.set("Metric System — completed", "ok")

            messagebox.showinfo(
                "Done",
                "\n".join([
                    "Metric System filter complete.\n",
                    f"Files processed        : {len(xliff_paths)}",
                    f"Segments evaluated     : {grand_total}",
                    f"───────────────────────────────────",
                    f"QA_Report  (errors)    : {grand_errors}",
                    f"Exclusions (non-errors): {grand_excl}",
                    f"  of which low-conf FP : {grand_low_conf}  ← yellow",
                    f"  of which clean       : {grand_clean}  ← green",
                    f"───────────────────────────────────",
                    f"QA + Excl = Total      : "
                    f"{grand_errors + grand_excl} / {grand_total}",
                    f"Confidence threshold   : {FP_CONFIDENCE_THRESHOLD:.2f}",
                    f"\nOutput folder:\n{out_dir}",
                ]),
            )

        except Exception as exc:
            self._status_bar.set(f"Error: {exc}", "error")
            self._log(f"Fatal error: {exc}", "error")
            messagebox.showerror("Error", str(exc))

    # ── helpers ───────────────────────────────────────────────────────────────
    def _stat_label(self, parent: tk.Frame, label: str, value: str) -> tk.Label:
        cell = tk.Frame(parent, bg=BG_CARD)
        cell.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=6, pady=6)
        tk.Label(cell, text=label, bg=BG_CARD,
                 fg=TEXT_MUTED, font=("Segoe UI", 7)).pack()
        lbl = tk.Label(cell, text=value, bg=BG_CARD,
                       fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"))
        lbl.pack()
        return lbl

    def _update_stats(
        self, total: int, errors: int, excl: int, clean: int
    ) -> None:
        self._lbl_total.config(text=str(total))
        self._lbl_real.config(text=str(errors), fg=DANGER if errors else SUCCESS)
        self._lbl_excl.config(text=str(excl),   fg=WARNING)
        self._lbl_clean.config(text=str(clean),  fg=SUCCESS)

    def _build_agent_cards(self, parent: tk.Frame) -> None:
        cards_frame = tk.Frame(parent, bg=BG_DARK)
        cards_frame.pack(fill=tk.X, pady=2)
        for title, desc, color, icon, side_pad in [
            ("Agent 1", "Claude — XLIFF Segment Analysis", "#c97bf7", "🧠", (0, 4)),
            ("Agent 2", "Claude — FP / Error Evaluation",  "#c97bf7", "🧠", (4, 0)),
        ]:
            card  = tk.Frame(cards_frame, bg=BG_CARD)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=side_pad)
            strip = tk.Frame(card, bg=color, width=3)
            strip.pack(side=tk.LEFT, fill=tk.Y)
            strip.pack_propagate(False)
            inner = tk.Frame(card, bg=BG_CARD, padx=10, pady=8)
            inner.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            tk.Label(
                inner, text=f"{icon} {title}",
                bg=BG_CARD, fg=color,
                font=("Segoe UI", 8, "bold"),
            ).pack(anchor="w")
            tk.Label(
                inner, text=desc,
                bg=BG_CARD, fg=TEXT_SECONDARY,
                font=("Segoe UI", 8),
            ).pack(anchor="w")

    def _log(self, text: str, level: str = "info") -> None:
        stamp = datetime.now().strftime("%H:%M:%S")
        if level == "info":
            tl = text.lower()
            if "error" in tl or "fail" in tl:
                level = "error"
            elif "agent" in tl or "claude" in tl:
                level = "agent"
            elif (
                "complet" in tl or "finish" in tl
                or "done" in tl or "saved" in tl
            ):
                level = "success"
            elif "warning" in tl or "fallback" in tl or "skip" in tl:
                level = "warning"

        self.log_box.config(state=tk.NORMAL)
        self.log_box.insert(tk.END, f"[{stamp}]  ", "time")
        self.log_box.insert(tk.END, f"{text}\n", level)
        self.log_box.see(tk.END)
        self.log_box.config(state=tk.DISABLED)
        self.update_idletasks()

    def _clear_log(self) -> None:
        self.log_box.config(state=tk.NORMAL)
        self.log_box.delete("1.0", tk.END)
        self.log_box.config(state=tk.DISABLED)

    def _set_progress_abs(self, pct: int, done: int, total: int) -> None:
        self._progress["value"] = max(0, min(100, pct))
        self._progress_lbl.config(text=f"{done} / {total}  ({pct}%)")
        self.update_idletasks()

    def _set_progress(self, done: int, total: int) -> None:
        if total:
            self._set_progress_abs(int(done / total * 100), done, total)

    def _run(self) -> None:
        threading.Thread(target=self._run_inner, daemon=True).start()


# ═══════════════════════════════════════════════════════════════════════════════
#  Main Application
# ═══════════════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1100x860")
        self.minsize(900, 700)
        self.configure(bg=BG_DARK)

        try:
            self.tk.call("tk", "scaling", 1.25)
        except Exception:
            pass

        self._build_ui()

    def _build_ui(self) -> None:

        # ── header ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=BG_PANEL, height=56)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(
            header,
            text="⬡  Indeed Proto QA Reviewer",
            bg=BG_PANEL, fg=TEXT_PRIMARY,
            font=("Segoe UI", 13, "bold"),
        ).pack(side=tk.LEFT, padx=18, pady=14)

        tk.Label(
            header,
            text="2-Agent Agentic Workflow  |  Claude Opus + Claude Haiku",
            bg=BG_PANEL, fg=TEXT_MUTED,
            font=("Segoe UI", 8),
        ).pack(side=tk.RIGHT, padx=18)

        tk.Frame(self, bg=ACCENT, height=2).pack(fill=tk.X)

        # ── status bar — packed to bottom first so it is always visible ───────
        self._status_bar = StatusBar(self)
        self._status_bar.pack(fill=tk.X, side=tk.BOTTOM)

        tk.Frame(self, bg=ACCENT, height=2).pack(fill=tk.X, side=tk.BOTTOM)

        # ── notebook ──────────────────────────────────────────────────────────
        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            "QA.TNotebook",
            background=BG_DARK,
            borderwidth=0,
            tabmargins=[0, 0, 0, 0],
        )
        style.configure(
            "QA.TNotebook.Tab",
            background=BG_PANEL,
            foreground=TEXT_SECONDARY,
            font=("Segoe UI", 9, "bold"),
            padding=[18, 8],
            borderwidth=0,
            focuscolor=BG_DARK,
        )
        style.map(
            "QA.TNotebook.Tab",
            background=[
                ("selected", BG_DARK),
                ("active",   BG_CARD),
            ],
            foreground=[
                ("selected", TEXT_PRIMARY),
                ("active",   TEXT_PRIMARY),
            ],
            expand=[("selected", [0, 0, 0, 2])],
        )

        self._notebook = ttk.Notebook(self, style="QA.TNotebook")
        self._notebook.pack(fill=tk.BOTH, expand=True)

        # Tab 1 — QA Reviewer (shell, disabled)
        self._tab1 = Tab1Frame(self._notebook, self._status_bar)
        self._notebook.add(self._tab1, text="  Indeed Proto QA Reviewer  ")

        # Tab 2 — Metric System (active)
        self._tab2 = Tab2Frame(self._notebook, self._status_bar)
        self._notebook.add(self._tab2, text="  METRIC SYSTEM  ")

        # Tab 3 — DNT r3 (placeholder)
        self._tab3 = self._build_placeholder_tab("DNT (r3)")
        self._notebook.add(self._tab3, text="  DNT (r3)  ")

        # Open on Tab 2 by default
        self._notebook.select(1)

    def _build_placeholder_tab(self, title: str) -> tk.Frame:
        frame = tk.Frame(self._notebook, bg=BG_DARK)

        tk.Frame(frame, bg=BG_DARK, height=80).pack()

        tk.Label(
            frame, text="⬡",
            bg=BG_DARK, fg=ACCENT_DIM,
            font=("Segoe UI", 48),
        ).pack()

        tk.Label(
            frame, text=title,
            bg=BG_DARK, fg=TEXT_PRIMARY,
            font=("Segoe UI", 18, "bold"),
        ).pack(pady=(12, 4))

        tk.Label(
            frame,
            text="This module is under development.",
            bg=BG_DARK, fg=TEXT_SECONDARY,
            font=("Segoe UI", 10),
        ).pack()

        tk.Label(
            frame, text="Coming soon.",
            bg=BG_DARK, fg=TEXT_MUTED,
            font=("Segoe UI", 9),
        ).pack(pady=(2, 0))

        tk.Frame(frame, bg=BG_DARK, height=24).pack()
        tk.Frame(frame, bg=ACCENT_DIM, height=2, width=120).pack()

        return frame


# ═══════════════════════════════════════════════════════════════════════════════
#  Entry point
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()